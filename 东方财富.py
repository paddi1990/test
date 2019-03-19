import requests
import re
from lxml import etree
import json
import openpyxl
from openpyxl import Workbook
def writeDataToExcel():
    token = '4f1862fc3b5e77c150a2b985b12db0fd'
    # 存放分页数据
    data_list = []
    # 表格标题
    title = []
    # 标题数量
    title_size = 26
    # 生成标题
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    for i in range(title_size):
        title.append('line%s' % i)
        worksheet.cell(1,i+1,title[i])
    # 迭代分页数据

    row=2
    for page in range(1, 3):
        response = requests.get(
            'http://nufm.dfcfw.com/EM_Finance2014NumericApplication/JS.aspx?type=CT&token=%s&sty=FCOIATC&cmd=C._A&st=(ChangePercent)&p=%d&ps=20' % (
                token, page))
        # 请求成功解析数据
        if response.status_code == 200:
            print("开始解析第%d页的数据" % page)
            # 页面解码
            datas = response.content.decode(response.apparent_encoding)
            if '"' in datas:
                datas = datas.split('"')
                for data in datas:
                    # 按逗号切割数据
                    if ',' in data and len(data.split(',')) == title_size:
                        line_data = data.split(',')
                        # 每一行的数据
                        data_json = {}
                        for index in range(title_size):
                            worksheet.cell(row,index+1,line_data[index])
                            data_json[title[index]] = line_data[index]
                        print(data_json)
                        data_list.append(data_json)
                        row += 1
        else:
            print('第%d页数据请求失败' % page)
    with open('data.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(data_list, indent=1, ensure_ascii=False))
    workbook.save('data.xlsx')

if __name__ == '__main__':
    writeDataToExcel()
