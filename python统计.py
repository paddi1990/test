import os
import math
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

result = {}


def results(key, value):
    result[key] = value
    print('%s：%s' % (key, str(value)))

#置信度
s={"95%":1.96,"90%":1.645,"99.7%":2.58}

# filepath=加载数据报表文件路径
# limits=分组界限
def question1(filepath, limits=[0, 1500, 2000, 2500, 3000], target='直方图.xlsx'):
    if os.path.isfile(filepath):

        wb = load_workbook(filepath)

        ws = wb.worksheets[0]

        # 初始化分组
        groups = {}
        for index in range(1, len(limits)):
            groups[str(limits[index - 1]) + "-" + str(limits[index])] = 0

        # 工薪总和
        payrolls = 0
        # 分组
        payroll_list = []
        for row in range(2, ws.max_row + 1):
            payroll = ws.cell(row=row, column=2).value
            payrolls += payroll
            payroll_list.append(payroll)
            for index in range(1, len(limits)):
                if index < len(limits) and payroll > limits[index - 1] and payroll <= limits[index]:
                    groups[str(limits[index - 1]) + "-" + str(limits[index])] += 1
                    break

        #样本均值
        avg = payrolls /len(payroll_list)
        results('样本均值', format(avg, '0.2f'))

        # 抽样平均误差
        sampling_avg_error = 0
        for payroll in payroll_list:
            sampling_avg_error = sampling_avg_error + math.pow(payroll - avg, 2)
        sampling_avg_error = math.sqrt(sampling_avg_error / len(payroll_list))
        results('抽样平均误差', format(sampling_avg_error, '0.2f'))

        # #抽样平均误差
        # sampling_avg_error=sampling_avg_error/math.sqrt(len(payroll_list))
        # results('抽样平均误差', format(sampling_avg_error, '0.2f'))

        for i in s:
            result[i]={}
            # 置信区间
            result[i]['置信区间']='('+format(avg-s[i]*(sampling_avg_error/math.sqrt(len(payroll_list))),'0.2f')+','+format(avg+s[i]*(sampling_avg_error/math.sqrt(len(payroll_list))),'0.2f')+')'
            #极限误差
            result[i]['抽样极限误差']=format(sampling_avg_error*s[i], '0.2f')
            #总体比重区间
            for index in range(1, len(limits)):
                group = str(limits[index - 1]) + "-" + str(limits[index])
                p=groups[group]/len(payroll_list)
                j=s[i] *math.sqrt(p*(1-p)/len(payroll_list))
                result[i][group+'总体比重区间'] ='('+format((p-j),'0.2f')+','+format((p+j),'0.2f')+')'


        # 写入分组数据
        for index in range(1, len(limits)):
            ws.cell(index, 3, str(limits[index - 1]) + "-" + str(limits[index]))
            ws.cell(index, 4, groups[str(limits[index - 1]) + "-" + str(limits[index])])

        # 绘制图表
        chart1 = BarChart()
        chart1.title = "工薪分布情况"
        # y轴标题
        chart1.y_axis.title = '人数'
        # x轴标题
        chart1.x_axis.title = '工薪区间'
        # 分组数据
        data = Reference(ws, min_col=4, min_row=1, max_row=len(limits))
        # 分组标题
        titles = Reference(ws, min_col=3, min_row=1, max_row=len(limits))
        chart1.add_data(data)
        chart1.set_categories(titles)
        # 图表位置
        ws.add_chart(chart1, "E1")
        # 生成文件路径
        wb.save(target)


if __name__ == '__main__':
    question1('1.xlsx')
