import argparse
import json
import math
import os
import random
import re
import time

from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference

# 记录数值
result = {}

confidence = {"95%": 1.96, "90%": 1.645, "99.7%": 2.58}
default_confidence_key = ['90%', '95%']
limits = [0, 1500, 2000, 2500, 3000]
desc = '直方图.xlsx'


def match(path):
    return re.fullmatch('[a-zA-Z]{1}:([/\\\]\w)+\\.xlsx|\w+\\.xlsx', path) is not None


# src=输入Excel文件路径
# limits=分组界限
# confidence=置信度
# desc=输出Excel文件路径
def main(src, limits=limits, value=default_confidence_key, desc=desc):
    if not match(src) or not match(desc):
        print('请检查：\n输入文件路径：%s\n输出文件路径：%s' % (src, desc))
    elif os.path.isfile(src):
        try:
            wb = load_workbook(src)
        except BaseException as err:
            print('文件已损坏')
        else:
            ws = wb.worksheets[0]
            limits.sort()
            # 初始化分组
            groups = {}
            for index in range(1, len(limits)):
                groups[str(limits[index - 1]) + "-" + str(limits[index])] = 0

            # 工薪总和
            payrolls = 0
            # 分组
            payroll_list = []
            for column in range(2, ws.max_row + 1):
                payroll = ws.cell(row=column, column=2).value
                payrolls += payroll
                payroll_list.append(payroll)
                for index in range(1, len(limits)):
                    if index < len(limits) and payroll > limits[index - 1] and payroll <= limits[index]:
                        groups[str(limits[index - 1]) + "-" + str(limits[index])] += 1
                        break

            # 样本均值
            avg = payrolls / len(payroll_list)
            result['样本均值'] = format(avg, '0.2f')

            # 抽样平均误差
            sampling_avg_error = 0
            for payroll in payroll_list:
                sampling_avg_error = sampling_avg_error + math.pow(payroll - avg, 2)
            sampling_avg_error = math.sqrt(sampling_avg_error / len(payroll_list))
            result['抽样平均误差'] = format(sampling_avg_error, '0.2f')

            # #抽样平均误差
            # sampling_avg_error=sampling_avg_error/math.sqrt(len(payroll_list))
            # results('抽样平均误差', format(sampling_avg_error, '0.2f'))

            for i in value:
                result[i] = {}
                # 置信区间
                if i not in confidence:
                    continue
                result[i]['置信区间'] = '(' + format(
                    avg - confidence[i] * (sampling_avg_error / math.sqrt(len(payroll_list))),'0.2f') + ',' + format(avg + confidence[i] * (sampling_avg_error / math.sqrt(len(payroll_list))), '0.2f') + ')'
                # 极限误差
                result[i]['抽样极限误差'] = format(sampling_avg_error * confidence[i], '0.2f')
                # 总体比重区间
                for index in range(1, len(limits)):
                    group = str(limits[index - 1]) + "-" + str(limits[index])
                    p = groups[group] / len(payroll_list)
                    j = confidence[i] * math.sqrt(p * (1 - p) / len(payroll_list))
                    result[i][group + '总体比重区间'] = '(' + format((p - j), '0.2f') + ',' + format((p + j), '0.2f') + ')'

            with open('数据.json', 'w', encoding='utf-8') as f:
                json_str = json.dumps(result, ensure_ascii=False, indent=1)
                print(json_str)
                f.write(json_str)

            # 写入分组数据

            column = ws.max_column + 1
            ws.cell(1, column, '工薪区间')
            ws.cell(1, column + 1, '人数')
            for index in range(1, len(limits)):
                ws.cell(index + 1, column, str(limits[index - 1]) + "-" + str(limits[index]))
                ws.cell(index + 1, column + 1, groups[str(limits[index - 1]) + "-" + str(limits[index])])

            # 绘制图表
            chart1 = BarChart()
            chart1.title = "工薪分布情况"
            # y轴标题
            chart1.y_axis.title = '人数'
            # x轴标题
            chart1.x_axis.title = '工薪区间'
            # 分组数据
            data = Reference(ws, min_col=ws.max_column, min_row=2, max_row=len(limits))
            # 分组标题
            titles = Reference(ws, min_col=ws.max_column - 1, min_row=2, max_row=len(limits))
            chart1.add_data(data)
            chart1.set_categories(titles)
            # 图表位置
            ws.add_chart(chart1, ws.cell(1, ws.max_column + 1).coordinate)
            # 生成文件路径
            wb.save(desc)
            print('图表生成到：%s' % desc)
    else:
        print('文件：%s不存在', src)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='统计工薪小程序')
    parser.add_argument('-s', '--src', type=str, help='测试样本路径')
    parser.add_argument('-n', '--num', type=int, help='随机样本数量')
    parser.add_argument('-min', '--min', type=int, help='生成样本最小值',default=min(limits))
    parser.add_argument('-max', '--max', type=int, help='生成样本最大值',default=max(limits))
    parser.add_argument("-l", '--limit', type=int, nargs='+', help='样本界限', default=limits)
    parser.add_argument('-d', '--desc', type=str, help='图表生成路径', default=desc)
    parser.add_argument('-c', '--confidence', type=int, nargs='+', help='置信度', choices=confidence.keys(),
                        default=default_confidence_key)
    args = vars(parser.parse_args())

    if args['src'] is not None and args['src'] != '':
        filepath = args['src']
        main(filepath, limits=args['limit'], desc=args['desc'], value=args['confidence'])
    elif args['num'] is not None and args['min']<args['max']:
        wb = workbook = Workbook()
        ws = wb.worksheets[0]
        ws.cell(1, 1, '序号')
        ws.cell(1, 2, '工薪')
        for index in range(2, args['num'] + 2):
            ws.cell(index, 1, index)
            ws.cell(index, 2, random.randint(args['min'], args['max']))
        filepath = '测试样本%d.xlsx' % int(time.time())
        wb.save(filepath)
        main(filepath, limits=args['limit'], desc=args['desc'], value=args['confidence'])
    else:
        parser.print_help()
