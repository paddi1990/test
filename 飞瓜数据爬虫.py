import json
import os
import sys
import time
import requests
import re
from lxml import etree
import argparse
# 解析图表json数据
from openpyxl import load_workbook, Workbook


def write(file, content):
    with open(file, 'w', encoding='utf-8') as f:
        f.write(content)


class FeiGuaData:
    # 请求头
    headers = {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 8.1.0; ALP-AL00 Build/HUAWEIALP-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/63.0.3239.83 Mobile Safari/537.36 T7/10.13 baiduboxapp/10.13.0.11 (Baidu; P1 8.1.0)',
        'Content-Type': 'application/json;charset=UTF-8',
        'Referer': 'https://dsp.xiguaji.com/',
        'Origin': 'https: // dsp.xiguaji.com',

    }

    # 数据概览标题头
    DETAIL_TITLE = ['昵称', '抖音号', '地区', '年龄', '分类', '粉丝数', '昨日排名打败', '近一周排名打败', '总点赞', '平均点赞', '总评论', '平均评论', '总分享',
                    '平均分享',
                    '近30天粉丝增量', '近90天视频发布']

    # 抖音热视频{"标题"："宽度"}
    AWEME_TITLE = {'传播指数': 8, '视频标题': 100, '播主': 30, '视频热词': 50, '视频时长': 8, '点赞数': 8, '评论数': 8, '分享数': 8, '发布时间': 17,
                   '数据更新时间': 17, '标签': 8, '视频链接': 90}

    LIST_TITLE = ['粉丝趋势90天总量', '粉丝趋势90天增量', '评论趋势总量', '评论趋势增量', '点赞趋势总量', '点赞趋势增量']

    # 年龄段
    AGE_TITLE = ['6-17', '18-24', '25-30', '31-35', '36-40', '41+']

    # 星座
    HOROSCOPE_TITLE = ['金牛座', '天秤座', '双鱼座', '摩羯座', '巨蟹座', '天蝎座', '狮子座', '白羊座', '双子座', '射手座', '水瓶座', '处女座']

    # 标签
    TAGS = ['', '网红美女', '网红帅哥', '搞笑', '情感', '剧情', '美食', '美妆', '穿搭', '明星', '影视娱乐', '游戏', '宠物',
            '音乐', '舞蹈', '萌娃', '生活', '体育', '旅行', '动漫', '创意', '时尚', '母婴育儿', '教育', '职场教育', '汽车', '家居',
            '科技', '摄影教学', '地方', '知识资讯类', '办公软件', '文学艺术']

    # 点赞数
    LIKES = ['1', '1-5', '5-10', '10-50', '50-100', '100-200', '200']

    # 小时
    HOURS = ['6', '12', '24', str(24 * 3), str(24 * 7), str(24 * 15), str(24 * 30)]

    # 视频时长
    DURATION = ['0', '15', '15-30', '30-60', '60']

    # 排序规则
    SORT = [0, 1, 2, 3]

    # 性别
    SEX_TITLE = ['男', '女']

    # 数据保存类型
    EXCEL = "EXCEL"

    @staticmethod
    def getJson(html, keyword, start, end):
        if keyword in html:
            html = re.findall(keyword + ' = .*', html)
            if len(html) > 0 and '[' in start and ']' in end:
                html = html[0]
                start = html.index('[')
                end = html.index(']')
                return json.loads(html[start:end + 1])

    # 提取图表左侧数据概览的数据
    @staticmethod
    def insertData(data_json, tree, class_name, tag=''):
        key = tree.xpath('//div[@class="' + class_name + '"]/' + tag + 'text()')
        value = tree.xpath('//div[@class="' + class_name + '"]/span/text()')
        if len(key) == 1 and len(value) == 1:
            data_json[key[0].strip()] = value[0].strip()

    @staticmethod
    def insertArea(data_json, areas, tag):
        for area in areas:
            if len(area.xpath('./td')) == 2:
                data_json['地域分布'][tag][area.xpath('./td/text()')[0]] = area.xpath('./td/text()')[1]

    @staticmethod
    def initSheet(workbook, *titles):
        for title in titles:
            workbook.create_sheet(title=title)
            workbook[title].cell(2, 1, '抖音号')
            workbook[title].column_dimensions[workbook[title].cell(2, 1, ).column_letter].width = 12
            if title == '性别分布':
                for index in range(len(FeiGuaData.SEX_TITLE)):
                    workbook[title].cell(1, index + 2, FeiGuaData.SEX_TITLE[index])
            elif title == '年龄分布':
                for index in range(len(FeiGuaData.AGE_TITLE)):
                    workbook[title].cell(1, index + 2, FeiGuaData.AGE_TITLE[index])
            elif title == '星座分布':
                for index in range(len(FeiGuaData.HOROSCOPE_TITLE)):
                    workbook[title].cell(1, index + 2, FeiGuaData.HOROSCOPE_TITLE[index])
            elif title not in ['地域分布', '近期10个作品表现']:
                workbook[title].cell(1, 2, '日期')

    @staticmethod
    def setdata_1(data, workbook, title, titles, row, id, key=None):
        if title in data.keys() and len(data[title]) > 0:
            worksheet = workbook[title]
            worksheet.cell(row + 1, 1, id)
            for index in range(len(titles)):
                if titles[index] in data[title]:
                    worksheet.cell(row + 1, index + 2, data[title][titles[index]])
                elif isinstance(data[title], list) and key is not None:
                    for i in range(len(data[title])):
                        worksheet.cell(row + 1, i + 2, data[title][i][key])

    # 数据文件名
    AWEME_FILE_NAME = 'aweme'

    @staticmethod
    def saveAwemeData(datas, type=EXCEL):
        if FeiGuaData.EXCEL == type:
            flag = not os.path.isfile('%s.xlsx' % FeiGuaData.AWEME_FILE_NAME)
            if flag:
                # 创建工作簿
                workbook = Workbook()
                # 创建工作表，初始化工作表表头
                for index in range(len(FeiGuaData.AWEME_TITLE.keys())):
                    workbook.worksheets[0].cell(1, index + 1, list(FeiGuaData.AWEME_TITLE.keys())[index])
                    workbook.worksheets[0].column_dimensions[
                        workbook.worksheets[0].cell(1, index + 1).column_letter].width = FeiGuaData.AWEME_TITLE[
                        list(FeiGuaData.AWEME_TITLE.keys())[index]]
            else:
                workbook = load_workbook('%s.xlsx' % FeiGuaData.AWEME_FILE_NAME)
            worksheet = workbook.worksheets[0]
            for data in datas:
                row = worksheet.max_row + 1
                for index in range(len(FeiGuaData.AWEME_TITLE.keys())):
                    if list(FeiGuaData.AWEME_TITLE.keys())[index] in data:
                        worksheet.cell(row, index + 1, str(data[list(FeiGuaData.AWEME_TITLE.keys())[index]]))
            workbook.save('%s.xlsx' % FeiGuaData.AWEME_FILE_NAME)
        else:
            write('%s.json' % FeiGuaData.AWEME_FILE_NAME, str(datas))

    # 数据文件名
    DATA_FILENAME = "data"

    @staticmethod
    def saveData(datas, type=EXCEL):
        if FeiGuaData.EXCEL == type:
            flag = not os.path.isfile('%s.xlsx' % FeiGuaData.DATA_FILENAME)
            if flag:
                # 创建工作簿
                workbook = Workbook()
                # 创建工作表，初始化工作表表头
                workbook.worksheets[0].title = '数据概览'
                for index in range(len(FeiGuaData.DETAIL_TITLE)):
                    workbook['数据概览'].cell(1, index + 1, FeiGuaData.DETAIL_TITLE[index])
                FeiGuaData.initSheet(workbook, '粉丝趋势90天总量', '粉丝趋势90天增量', '性别分布', '年龄分布', '地域分布', '星座分布', '评论趋势总量',
                                     '评论趋势增量', '点赞趋势总量', '点赞趋势增量', '近期10个作品表现')

            else:
                workbook = load_workbook('%s.xlsx' % FeiGuaData.DATA_FILENAME)

            # 往单元格内写入内容

            str_row = 2
            for data in datas:
                str_column = 1
                worksheet = workbook['数据概览']
                if '抖音号' in data:
                    id = data['抖音号']
                row = (str_row if flag else worksheet.max_row + 1)
                for title in FeiGuaData.DETAIL_TITLE:
                    if title in data.keys():
                        worksheet.cell(row, str_column, data[title])
                    str_column += 1

                for title in FeiGuaData.LIST_TITLE:
                    worksheet = workbook[title]
                    worksheet.cell(str_row + 1 if flag else worksheet.max_row + 1, 1, id)
                    str_column = 3
                    if title in data and not data[title] is None and 'None' != data[title]:
                        row = (str_row - 1 if flag else worksheet.max_row + 1)
                        _row = (str_row + 1 if flag else worksheet.max_row)
                        for sum in data[title]:
                            if flag and str_row == 2:
                                worksheet.cell(row, str_column,
                                               time.strftime("%Y/%m/%d", time.localtime(sum['x'] / 1000)))
                                worksheet.column_dimensions[worksheet.cell(row, str_column).column_letter].width = 12
                            worksheet.cell(_row, str_column, sum['y'])
                            str_column += 1

                FeiGuaData.setdata_1(data, workbook, '星座分布', FeiGuaData.HOROSCOPE_TITLE,
                                     str_row if flag else worksheet.max_row + 1, id)
                FeiGuaData.setdata_1(data, workbook, '性别分布', FeiGuaData.SEX_TITLE,
                                     str_row if flag else worksheet.max_row + 1, id)
                FeiGuaData.setdata_1(data, workbook, '年龄分布', FeiGuaData.AGE_TITLE,
                                     str_row if flag else worksheet.max_row + 1, id, 'count')

                worksheet = workbook['地域分布']

                row = (str_row + 1 if flag else worksheet.max_row + 1)
                worksheet.cell(row, 1, id)
                if '地域分布' in data:
                    index = 2
                    for key in data['地域分布']['省份']:
                        worksheet.cell(row, index, key + data['地域分布']['省份'][key])
                        # 调整列宽
                        worksheet.column_dimensions[worksheet.cell(row, index).column_letter].width = 12
                        index += 1
                    for key in data['地域分布']['城市']:
                        worksheet.cell(row, index, key + data['地域分布']['城市'][key])
                        worksheet.column_dimensions[worksheet.cell(row, index).column_letter].width = 12
                        index += 1

                worksheet = workbook['近期10个作品表现']
                worksheet.cell(row, 1, id)
                if '近期10个作品表现' in data:
                    for index in range(len(data['近期10个作品表现']['日期'])):
                        worksheet.cell(row, index + 2, '日期：%s,点赞量：%d,评论量：%d' % (
                            data['近期10个作品表现']['日期'][index], data['近期10个作品表现']['点赞量'][index],
                            data['近期10个作品表现']['评论量'][index]))
                        worksheet.column_dimensions[worksheet.cell(row, index + 1).column_letter].width = 35
                str_row += 1
            workbook.save('%s.xlsx' % FeiGuaData.DATA_FILENAME)
        else:
            write('%s.json' % FeiGuaData.DATA_FILENAME, str(datas))

    # 系统登录
    @staticmethod
    def login(tel, pwd):
        if tel == None or pwd == None:
            raise BaseException("账号或密码不能为空")
        else:
            response = requests.post('https://dsp.xiguaji.com/Login/Login', data=json.dumps({"tel": tel, "pwd": pwd}),
                                     headers=FeiGuaData.headers)
            result = response.content.decode(response.apparent_encoding)
            if response.status_code == 200 and '登陆成功' in result:
                return response.cookies
            else:
                print(response.content.decode(response.apparent_encoding))
                raise BaseException("登陆失败，请检查账号和密码")

    # 关键词搜索抖音号
    @staticmethod
    def searchKeyword(cookies, keyword):
        response = requests.get('https://dsp.xiguaji.com/Blogger/Search?keyword=%s' % keyword,
                                headers=FeiGuaData.headers,
                                cookies=cookies)
        if response.status_code == 200:
            return response
        else:
            print(response.status_code)
            print(response.content.decode('utf-8'))
            raise BaseException("关键词搜索异常")

    # 解析详情页数据
    @staticmethod
    def readDetailData(response, cookies):
        html = etree.HTML(response.content.decode('utf-8'), etree.HTMLParser())
        hrefs = html.xpath('//li/a[@class="btns-unified"]/@href')
        # 获取搜索结果页简介数据
        data_list = []
        for href in hrefs:
            link = 'https://dsp.xiguaji.com' + href[1:]
            print("开始解析图表链接：%s" % link)
            response = requests.get(link, headers=FeiGuaData.headers,
                                    cookies=cookies)
            if response.status_code == 200:
                data_json = {}
                html = response.content.decode('utf-8')
                if '抱歉，用户权限不足' in html:
                    raise BaseException("免费版每日仅能查阅10次播主详情")
                tree = etree.HTML(html)
                # 昵称
                if len(tree.xpath('//div[@class="nickname v-tag"]')) == 1:
                    data_json['昵称'] = tree.xpath('//div[@class="nickname v-tag"]')[0].text.strip()
                key = tree.xpath('//div[@class="info"]//li/text()')
                value = tree.xpath('//div[@class="info"]//span/text()')
                # 抖音好/地区/年龄/分类
                _key = []
                for index in range(len(key)):
                    if key[index].strip() == '':
                        continue
                    _key.append(key[index].strip())
                for index in range(len(value)):
                    data_json[_key[index].replace('：', '')] = value[index].strip()
                # 粉丝数
                FeiGuaData.insertData(data_json, tree, 'col-sm-4 fans js-fans-tip', 'a/')
                # 作品数
                FeiGuaData.insertData(data_json, tree, 'col-sm-4 videos-count', 'a/')
                # 飞瓜指数
                FeiGuaData.insertData(data_json, tree, 'col-sm-4 xiagua-index', 'a/')
                # 数据概览数据
                other_infos = tree.xpath('//div[@class="owner-data"]//div[@class="row col2"]')
                for info in other_infos:
                    sub_infos = info.xpath('./div[@class="col-sm-6"]')
                    for sub_info in sub_infos:
                        if len(sub_info.xpath('./text()')) > 0 and len(sub_info.xpath('./span/text()')) > 0:
                            data_json[sub_info.xpath('./text()')[0].strip()] = \
                                sub_info.xpath('./span/text()')[0].strip()
                # 粉丝趋势
                data_json['粉丝趋势90天总量'] = FeiGuaData.getJson(html, 'fans90', '[', ']')
                data_json['粉丝趋势90天增量'] = FeiGuaData.getJson(html, 'fansInc90', '[', ']')
                # 性别分布
                left = tree.xpath('//div[@class="row gender-text"]/div[@class="col-sm-6 nopadding"]/text()')
                right = tree.xpath(
                    '//div[@class="row gender-text"]/div[@class="col-sm-6 nopadding text-right"]/text()')
                if len(left) == 2 and len(right) == 2:
                    data_json['性别分布'] = {left[0].strip(): left[1].strip(),
                                         right[0].strip(): right[1].strip()}
                # 年龄分布
                data_json['年龄分布'] = FeiGuaData.getJson(html, 'age_data', '[', ']')
                # 地域分布
                areas = tree.xpath('//div[@class="data-section location-section pt20"]//tbody')
                if len(areas) == 2:
                    data_json['地域分布'] = {'省份': {}, '城市': {}}
                    FeiGuaData.insertArea(data_json, areas[0], '省份')
                    FeiGuaData.insertArea(data_json, areas[1], '城市')
                # 星座分布
                zodiac_sections = tree.xpath('//div[@class="data-section zodiac-section pt20"]//li')
                data_json['星座分布'] = {}
                for zodiac_section in zodiac_sections:
                    if len(zodiac_section.xpath('./div[@class="zodiac-name"]')) == 1 and len(
                            zodiac_section.xpath('./div[@class="zodiac-percent"]')) == 1:
                        data_json['星座分布'][zodiac_section.xpath('./div[@class="zodiac-name"]//text()')[0]] = \
                            zodiac_section.xpath('./div[@class="zodiac-percent"]/text()')[0]
                # 评论趋势
                data_json['评论趋势总量'] = FeiGuaData.getJson(html, 'comment90', '[', ']')
                data_json['评论趋势增量'] = FeiGuaData.getJson(html, 'commentCountInc90', '[', ']')
                data_json['点赞趋势总量'] = FeiGuaData.getJson(html, 'likes90', '[', ']')
                data_json['点赞趋势增量'] = FeiGuaData.getJson(html, 'likeCountInc90', '[', ']')
                data_json['近期10个作品表现'] = {'日期': FeiGuaData.getJson(html, 'lineDatecode', '[', ']'),
                                          '点赞量': FeiGuaData.getJson(html, 'lineLike', '[', ']'),
                                          '评论量': FeiGuaData.getJson(html, 'lineComment', '[', ']')}
                print(data_json)
                data_list.append(data_json)
        return data_list

    @staticmethod
    def aweme(cookies, keyword='', tag='', likes='0', hours='24', duration='0', gender='0', age='0', province='0',
              city='0', sort='0',
              ispromotions='0', page='1'):
        link = 'https://dsp.xiguaji.com/Aweme/Search?keyword=%s&tag=%s&likes=%s&hours=%s&duration=%s&gender=%s&age=%s&province=%s&city=%s&sort=%s&ispromotions=%s&page=%s' % (
            keyword, tag, likes, hours, duration, gender, age, province, city, sort, ispromotions, page)
        print('link=%s' % link)
        response = requests.get(link, headers=FeiGuaData.headers, cookies=cookies)
        if response.status_code == 200:
            html = etree.HTML(response.content.decode('utf-8'), etree.HTMLParser())
            tr_list = html.xpath('//tr[@class="js-slider-aweme"]')
            data_list = []
            for tr in tr_list:
                data_json = {}
                # 传播指数
                if len(tr.xpath('.//span/text()')) > 0:
                    data_json['传播指数'] = tr.xpath('.//span/text()')[0].strip()
                item_title = tr.xpath('.//div[@class="item-title"]/a/text()')
                if len(item_title) == 2:
                    # 视频标题
                    data_json['视频标题'] = item_title[0].strip()
                    # 播主
                    data_json['播主'] = item_title[1].strip()
                # 视频热词
                if len(tr.xpath('.//div[@class="item-tag clearfix"]//a/text()')) > 0:
                    data_json['视频热词'] = '-'.join(tr.xpath('.//div[@class="item-tag clearfix"]//a/text()'))
                # 视频时长
                if len(tr.xpath('.//div[@class="item-times"]/text()')) == 2:
                    data_json['视频时长'] = tr.xpath('.//div[@class="item-times"]/text()')[1].strip()
                id = tr.xpath('.//i[@class="icon-details"]/../@data-id')
                awemeId = tr.xpath('.//i[@class="icon-details"]/../@data-awemeid')
                if len(id) == 1 and len(awemeId) == 1:
                    link = 'https://dsp.xiguaji.com/Aweme/Detail?id=%s&awemeId=%s&active=detail' % (id[0], awemeId[0])
                    response = requests.get(link, headers=FeiGuaData.headers, cookies=cookies)
                    if response.status_code == 200:
                        detail_tree = etree.HTML(response.content.decode('utf-8'), etree.HTMLParser())
                        # 点赞数
                        if len(detail_tree.xpath('//i[@class="v-icon-set like"]/../text()')) == 2:
                            data_json['点赞数'] = detail_tree.xpath('//i[@class="v-icon-set like"]/../text()')[1].strip()
                            # 评论数
                        if len(detail_tree.xpath('//i[@class="v-icon-set comments"]/../text()')) == 2:
                            data_json['评论数'] = detail_tree.xpath('//i[@class="v-icon-set comments"]/../text()')[
                                1].strip()
                            # 分享数
                        if len(detail_tree.xpath('//i[@class="v-icon-set reply"]/../text()')) == 2:
                            data_json['分享数'] = detail_tree.xpath('//i[@class="v-icon-set reply"]/../text()')[1].strip()

                        if len(detail_tree.xpath('//dl[@class="details-info-content"]/dd/text()')) == 1:
                            str = detail_tree.xpath('//dl[@class="details-info-content"]/dd/text()')[0]
                            times = re.findall('\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}', str)
                            # 发布时间
                            data_json['发布时间'] = times[0]
                            # 数据更新时间
                            data_json['数据更新时间'] = times[1]
                        # 标签
                        if len(detail_tree.xpath('//span[@class="details-tag"]/text()')) > 0:
                            data_json['标签'] = '-'.join(detail_tree.xpath('//span[@class="details-tag"]/text()'))
                        # 视频链接
                        if len(detail_tree.xpath('//a[@class="btn-play"]/@href')) == 1:
                            data_json['视频链接'] = detail_tree.xpath('//a[@class="btn-play"]/@href')[0]
                            print(data_json)
                            data_list.append(data_json)
            return data_list


# 控制台操作
def console():
    parser = argparse.ArgumentParser(description='爬取飞瓜数据命令行程序')
    subparsers = parser.add_subparsers(required=True, metavar='search/rank/Aweme 抖音号搜索/抖音排行榜/抖音热门视频')

    search = subparsers.add_parser('search', help='保存抖音号搜索数据')
    search.add_argument('-t', '--type', metavar='搜索类型', choices=['base', 'advanced'], required=True,
                        help='base=关键词搜索，advanced=高级搜索')

    rank = subparsers.add_parser('rank', help='保存抖音排行榜数据')
    rank.add_argument('-p', '--period', metavar='统计时间', choices=['day', 'week', 'month'],
                      help='day=日榜，week=周榜，month=月榜')

    Aweme = subparsers.add_parser('Aweme', help='保存抖音热门视频数据')

    # Aweme.add_argument('--tag', metavar='标签',
    #                    choices=FeiGuaData.TAGS,help=','.join(FeiGuaData.TAGS))

    # def aweme(cookies, keyword='', tag='', likes='0', hours=24, duration=0, gender=0, age=0, province=0, city=0, sort=0,
    #           ispromotions=0, page=1):

    Aweme.add_argument('--likes', metavar='点赞数', choices=FeiGuaData.LIKES, help='万,'.join(FeiGuaData.LIKES) + '万',
                       default='0')
    Aweme.add_argument('--hours', metavar='最近发布时间', choices=FeiGuaData.HOURS, help='小时,'.join(FeiGuaData.HOURS) + '小时',
                       default='24')
    Aweme.add_argument('--duration', metavar='视频时长', choices=FeiGuaData.DURATION,
                       help='0:不限,' + '秒,'.join(FeiGuaData.DURATION[1:]) + '秒', default='0')
    Aweme.add_argument('--sort', metavar='排序规则', choices=FeiGuaData.SORT, help='1:点赞最多,2:评论最多,3:分享最多', default='0')
    Aweme.add_argument('--ispromotions', metavar='商品橱窗', choices=[0, 1], help='1:已开通,0:未开通', default='0')
    Aweme.add_argument('--page', metavar='页数', type=int, default='0')

    parser.add_argument('-u', '--user', metavar='手机号', required=True, type=str)
    parser.add_argument('-p', '--password', metavar='密码', required=True, type=str)
    search.add_argument('-k', '--keyword', metavar='关键词', required=True, type=str)

    args = vars(parser.parse_args())

    cookies = FeiGuaData.login(args['user'], args['password'])
    if 'search' in sys.argv:
        if 'base' in sys.argv:
            print('关键词搜索')
            response = FeiGuaData.searchKeyword(cookies, args.keyword)
            datas = FeiGuaData.readDetailData(response, cookies)
            FeiGuaData.saveData(datas, FeiGuaData.EXCEL)
        elif 'advanced' in sys.argv:
            print('高级搜索')
            # response = FeiGuaData.searchKeyword(cookies, args.keyword)
            # datas = FeiGuaData.readDetailData(response, cookies)
    elif 'rank' in sys.argv:
        print('抖音排行榜')
    elif 'Aweme' in sys.argv:
        print('抖音热门视频')
        datas = FeiGuaData.aweme(cookies, keyword=args['keyword'] if 'keyword' in args else '', likes=args['likes'],
                                 hours=args['hours'], duration=args['duration'], sort=args['sort'],
                                 ispromotions=args['ispromotions'], page=args['page'])
        FeiGuaData.saveAwemeData(datas, FeiGuaData.EXCEL)



if __name__ == '__main__':
    console()
